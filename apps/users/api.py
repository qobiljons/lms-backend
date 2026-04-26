import io
import os

from django_filters import rest_framework as django_filters
from drf_yasg.utils import swagger_auto_schema
from rest_framework import generics, permissions, serializers as drf_serializers, status
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .permissions import IsAdmin
from .serializers import (
    AdminCreateUserSerializer,
    AdminSetPasswordSerializer,
    AdminUpdateUserSerializer,
    ChangePasswordSerializer,
    LoginSerializer,
    SignupSerializer,
    UpdateMeSerializer,
    UserProfileSerializer,
    UserSerializer,
)

def _get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        "refresh": str(refresh),
        "access": str(refresh.access_token),
    }

class LogoutRequestSerializer(drf_serializers.Serializer):
    refresh = drf_serializers.CharField(help_text="The refresh token to blacklist")

class SignupAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    authentication_classes = ()

    @swagger_auto_schema(request_body=SignupSerializer, responses={201: UserSerializer})
    def post(self, request):
        serializer = SignupSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        tokens = _get_tokens_for_user(user)
        return Response(
            {**UserSerializer(user).data, "tokens": tokens},
            status=status.HTTP_201_CREATED,
        )

class LoginAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    authentication_classes = ()

    @swagger_auto_schema(request_body=LoginSerializer, responses={200: UserSerializer})
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        tokens = _get_tokens_for_user(user)
        return Response({**UserSerializer(user).data, "tokens": tokens})

class UserPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = "page_size"
    max_page_size = 1000

class UserListAPIView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = (permissions.IsAuthenticated, IsAdmin)
    pagination_class = UserPagination
    filter_backends = (django_filters.DjangoFilterBackend, SearchFilter, OrderingFilter)
    filterset_fields = ("role", "is_active")
    search_fields = ("username", "email", "first_name", "last_name")
    ordering_fields = ("username", "email", "date_joined")
    ordering = ("-date_joined",)

    def get_queryset(self):
        from django.contrib.auth import get_user_model
        return get_user_model().objects.all()

class UserDetailAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsAdmin)

    def _get_user(self, username):
        from django.contrib.auth import get_user_model
        try:
            return get_user_model().objects.get(username=username)
        except get_user_model().DoesNotExist:
            return None

    @swagger_auto_schema(responses={200: UserSerializer})
    def get(self, request, username):
        user = self._get_user(username)
        if user is None:
            return Response({"detail": "not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(UserSerializer(user).data)

    @swagger_auto_schema(request_body=AdminUpdateUserSerializer, responses={200: UserSerializer})
    def patch(self, request, username):
        user = self._get_user(username)
        if user is None:
            return Response({"detail": "not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = AdminUpdateUserSerializer(user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserSerializer(user).data)

    def delete(self, request, username):
        user = self._get_user(username)
        if user is None:
            return Response({"detail": "not found"}, status=status.HTTP_404_NOT_FOUND)
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class AdminSetPasswordAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsAdmin)

    @swagger_auto_schema(request_body=AdminSetPasswordSerializer)
    def post(self, request, username):
        from django.contrib.auth import get_user_model
        try:
            user = get_user_model().objects.get(username=username)
        except get_user_model().DoesNotExist:
            return Response({"detail": "not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = AdminSetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user.set_password(serializer.validated_data["new_password"])
        user.save()
        return Response({"detail": "password updated"})

class AdminCreateUserAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsAdmin)

    @swagger_auto_schema(request_body=AdminCreateUserSerializer, responses={201: UserSerializer})
    def post(self, request):
        serializer = AdminCreateUserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)

class MeAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    @swagger_auto_schema(responses={200: UserSerializer})
    def get(self, request):
        return Response(UserSerializer(request.user).data)

    @swagger_auto_schema(request_body=UpdateMeSerializer, responses={200: UserSerializer})
    def patch(self, request):
        serializer = UpdateMeSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserSerializer(request.user).data)

class ProfileAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    @swagger_auto_schema(responses={200: UserProfileSerializer})
    def get(self, request):
        return Response(UserProfileSerializer(request.user.profile).data)

    @swagger_auto_schema(request_body=UserProfileSerializer, responses={200: UserProfileSerializer})
    def patch(self, request):
        serializer = UserProfileSerializer(
            request.user.profile, data=request.data, partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserProfileSerializer(request.user.profile).data)

class ChangePasswordAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    @swagger_auto_schema(request_body=ChangePasswordSerializer)
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        if not request.user.check_password(serializer.validated_data["old_password"]):
            return Response(
                {"old_password": ["incorrect password"]},
                status=status.HTTP_400_BAD_REQUEST,
            )
        request.user.set_password(serializer.validated_data["new_password"])
        request.user.save()
        return Response({"detail": "password updated"})

class LogoutAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    @swagger_auto_schema(request_body=LogoutRequestSerializer)
    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            if not refresh_token:
                return Response(
                    {"detail": "refresh token is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response({"success": True})
        except Exception:
            return Response(
                {"detail": "invalid token"},
                status=status.HTTP_400_BAD_REQUEST,
            )

class DashboardStatsAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    @staticmethod
    def _week_series(qs, date_field="created_at", weeks=7):
        """Return per-week counts for the last N weeks."""
        import datetime

        from django.db.models import Count
        from django.db.models.functions import TruncWeek
        from django.utils import timezone

        start = timezone.now() - datetime.timedelta(weeks=weeks)
        rows = (
            qs.filter(**{f"{date_field}__gte": start})
            .annotate(week=TruncWeek(date_field))
            .values("week")
            .annotate(count=Count("id"))
            .order_by("week")
        )
        lookup = {r["week"].strftime("%b %d"): r["count"] for r in rows}

        result = []
        for i in range(weeks):
            d = start + datetime.timedelta(weeks=i)
            d = d - datetime.timedelta(days=d.weekday())
            label = d.strftime("%b %d")
            result.append({"week": label, "count": lookup.get(label, 0)})
        return result

    @staticmethod
    def _daily_series(qs, date_field="session_date", days=14):
        """Return per-day counts for the last N days using a date field."""
        import datetime

        from django.db.models import Count
        from django.utils import timezone

        start = (timezone.now() - datetime.timedelta(days=days)).date()
        rows = (
            qs.filter(**{f"{date_field}__gte": start})
            .values(date_field)
            .annotate(count=Count("id"))
            .order_by(date_field)
        )
        lookup = {r[date_field].strftime("%b %d"): r["count"] for r in rows}

        result = []
        for i in range(days):
            d = start + datetime.timedelta(days=i)
            label = d.strftime("%b %d")
            result.append({"day": label, "count": lookup.get(label, 0)})
        return result

    @staticmethod
    def _attendance_status_by_session(sessions_qs, limit=10):
        """Return per-session attendance breakdown for bar chart."""
        recent = sessions_qs.order_by("-session_date")[:limit]
        result = []
        for s in reversed(recent):
            records = s.records.all()
            total = records.count()
            present = records.filter(status__in=["attended", "attended_online", "late"]).count()
            absent = records.filter(status="absent").count()
            result.append({
                "session": s.session_date.strftime("%b %d"),
                "present": present,
                "absent": absent,
                "total": total,
            })
        return result

    @staticmethod
    def _revenue_series(weeks=8):
        """Return per-week revenue for the last N weeks."""
        import datetime
        from decimal import Decimal

        from django.db.models import Sum
        from django.db.models.functions import TruncWeek
        from django.utils import timezone

        from apps.payments.models import Payment

        start = timezone.now() - datetime.timedelta(weeks=weeks)
        rows = (
            Payment.objects.filter(status="succeeded", created_at__gte=start)
            .annotate(week=TruncWeek("created_at"))
            .values("week")
            .annotate(total=Sum("amount"))
            .order_by("week")
        )
        lookup = {r["week"].strftime("%b %d"): float(r["total"] or 0) for r in rows}

        result = []
        for i in range(weeks):
            d = start + datetime.timedelta(weeks=i)
            d = d - datetime.timedelta(days=d.weekday())
            label = d.strftime("%b %d")
            result.append({"week": label, "revenue": lookup.get(label, 0)})
        return result

    def get(self, request):
        import datetime
        from decimal import Decimal

        from django.contrib.auth import get_user_model
        from django.db.models import Avg, Count, F, Q, Sum
        from django.utils import timezone

        from apps.attendance.models import AttendanceRecord, AttendanceSession
        from apps.courses.models import Course
        from apps.groups.models import Group
        from apps.homework.models import Homework, HomeworkSubmission
        from apps.lessons.models import Lesson
        from apps.payments.models import CoursePurchase, Payment

        User = get_user_model()
        user = request.user

        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        week_start = (now - datetime.timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0
        )

        if user.role == "admin":
            total_revenue = (
                Payment.objects.filter(status="succeeded").aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )
            monthly_revenue = (
                Payment.objects.filter(status="succeeded", created_at__gte=month_start)
                .aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )
            weekly_revenue = (
                Payment.objects.filter(status="succeeded", created_at__gte=week_start)
                .aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )

            succeeded_payments = Payment.objects.filter(status="succeeded")
            avg_payment = float(
                succeeded_payments.aggregate(avg=Avg("amount"))["avg"] or 0
            )

            total_records = AttendanceRecord.objects.count()
            present_records = AttendanceRecord.objects.filter(
                status__in=["attended", "attended_online", "late"]
            ).count()
            attendance_rate = round((present_records / total_records * 100), 1) if total_records > 0 else 0

            recent_payments = Payment.objects.filter(status="succeeded").select_related("user").order_by("-created_at")[:8]

            user_growth = self._week_series(User.objects.all(), "date_joined", weeks=8)
            revenue_trend = self._revenue_series(weeks=8)
            attendance_by_session = self._attendance_status_by_session(AttendanceSession.objects.all())

            att_status_dist = list(
                AttendanceRecord.objects.values("status")
                .annotate(count=Count("id"))
                .order_by("-count")
            )

            course_pop = []
            for c in Course.objects.all()[:6]:
                student_count = User.objects.filter(student_groups__courses=c).distinct().count()
                course_pop.append({"name": c.title[:20], "students": student_count})

            payment_status_dist = list(
                Payment.objects.values("status")
                .annotate(count=Count("id"))
                .order_by("-count")
            )

            revenue_by_course = []
            for cp_row in (
                CoursePurchase.objects.values("course__title")
                .annotate(total=Sum("amount"))
                .order_by("-total")[:6]
            ):
                revenue_by_course.append({
                    "name": (cp_row["course__title"] or "Unknown")[:20],
                    "revenue": float(cp_row["total"] or 0),
                })

            daily_sessions = self._daily_series(
                AttendanceSession.objects.all(), "session_date", days=14
            )

            attendance_rate_trend = []
            for i in range(8):
                wk_start = (week_start - datetime.timedelta(weeks=7 - i)).date()
                wk_end = wk_start + datetime.timedelta(weeks=1)
                wk_records = AttendanceRecord.objects.filter(
                    session__session_date__gte=wk_start,
                    session__session_date__lt=wk_end,
                )
                wk_total = wk_records.count()
                wk_present = wk_records.filter(
                    status__in=["attended", "attended_online", "late"]
                ).count()
                wk_rate = round((wk_present / wk_total * 100), 1) if wk_total > 0 else 0
                attendance_rate_trend.append({
                    "week": wk_start.strftime("%b %d"),
                    "rate": wk_rate,
                    "total": wk_total,
                    "present": wk_present,
                })

            group_attendance_breakdown = []
            for g in Group.objects.all().prefetch_related("students")[:10]:
                g_sessions = AttendanceSession.objects.filter(group=g)
                g_records = AttendanceRecord.objects.filter(session__in=g_sessions)
                g_total = g_records.count()
                if g_total == 0:
                    continue
                attended = g_records.filter(status="attended").count()
                online = g_records.filter(status="attended_online").count()
                late = g_records.filter(status="late").count()
                absent = g_records.filter(status="absent").count()
                excused = g_records.filter(status="excused").count()
                present_total = attended + online + late
                rate = round((present_total / g_total * 100), 1) if g_total > 0 else 0
                group_attendance_breakdown.append({
                    "name": g.name[:20],
                    "students": g.students.count(),
                    "sessions": g_sessions.count(),
                    "attended": attended,
                    "online": online,
                    "late": late,
                    "absent": absent,
                    "excused": excused,
                    "total": g_total,
                    "rate": rate,
                })
            group_attendance_breakdown.sort(key=lambda x: x["rate"], reverse=True)

            total_hw = Homework.objects.count()
            total_subs = HomeworkSubmission.objects.count()
            graded_subs = HomeworkSubmission.objects.filter(status="graded").count()
            pending_subs = HomeworkSubmission.objects.filter(status="submitted").count()
            draft_subs = HomeworkSubmission.objects.filter(status="draft").count()

            recent_users = [
                {
                    "username": u.username,
                    "name": f"{u.first_name} {u.last_name}".strip() or u.username,
                    "role": u.role,
                    "date": u.date_joined.strftime("%b %d, %Y"),
                }
                for u in User.objects.order_by("-date_joined")[:6]
            ]

            new_users_week = User.objects.filter(date_joined__gte=week_start).count()
            new_users_month = User.objects.filter(date_joined__gte=month_start).count()

            sessions_this_week = AttendanceSession.objects.filter(
                session_date__gte=week_start.date()
            ).count()

            avg_score_all = float(
                HomeworkSubmission.objects.filter(score__isnull=False)
                .aggregate(avg=Avg("score"))["avg"] or 0
            )

            total_payment_attempts = Payment.objects.count()
            payment_success_rate = (
                round((succeeded_payments.count() / total_payment_attempts * 100), 1)
                if total_payment_attempts > 0
                else 0
            )

            prev_month_end = month_start - datetime.timedelta(seconds=1)
            prev_month_start = prev_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            prev_month_revenue = float(
                Payment.objects.filter(
                    status="succeeded",
                    created_at__gte=prev_month_start,
                    created_at__lt=month_start,
                ).aggregate(total=Sum("amount"))["total"] or 0
            )
            current_month_revenue_float = float(monthly_revenue)
            if prev_month_revenue > 0:
                revenue_growth_pct = round(
                    ((current_month_revenue_float - prev_month_revenue) / prev_month_revenue) * 100, 1
                )
            elif current_month_revenue_float > 0:
                revenue_growth_pct = 100.0
            else:
                revenue_growth_pct = 0.0

            top_instructors_qs = (
                User.objects.filter(role="instructor")
                .annotate(
                    group_count=Count("instructed_groups", distinct=True),
                    student_count=Count("instructed_groups__students", distinct=True),
                )
                .order_by("-student_count", "-group_count")[:5]
            )
            top_instructors = [
                {
                    "username": u.username,
                    "name": f"{u.first_name} {u.last_name}".strip() or u.username,
                    "groups": u.group_count,
                    "students": u.student_count,
                }
                for u in top_instructors_qs
            ]

            courses_with_purchases = (
                CoursePurchase.objects.values("course").distinct().count()
            )
            total_courses_count = Course.objects.count()
            course_conversion_rate = (
                round((courses_with_purchases / total_courses_count * 100), 1)
                if total_courses_count > 0
                else 0
            )

            active_student_ids = set(
                HomeworkSubmission.objects.values_list("student_id", flat=True)
            ) | set(
                AttendanceRecord.objects.values_list("student_id", flat=True)
            )
            total_students = User.objects.filter(role="student").count()
            engagement_rate = (
                round((len(active_student_ids) / total_students * 100), 1)
                if total_students > 0
                else 0
            )

            activity_feed = []
            for u in User.objects.order_by("-date_joined")[:3]:
                activity_feed.append({
                    "type": "user",
                    "title": f"{u.first_name} {u.last_name}".strip() or u.username,
                    "subtitle": f"Joined as {u.role}",
                    "date": u.date_joined.strftime("%b %d, %Y"),
                    "timestamp": u.date_joined.timestamp(),
                })
            for p in Payment.objects.filter(status="succeeded").select_related("user").order_by("-created_at")[:3]:
                activity_feed.append({
                    "type": "payment",
                    "title": f"${p.amount} payment",
                    "subtitle": f"by @{p.user.username}",
                    "date": p.created_at.strftime("%b %d, %Y"),
                    "timestamp": p.created_at.timestamp(),
                })
            for s in HomeworkSubmission.objects.filter(status="submitted").select_related("student", "homework").order_by("-created_at")[:3]:
                activity_feed.append({
                    "type": "submission",
                    "title": s.homework.title[:35],
                    "subtitle": f"submitted by @{s.student.username}",
                    "date": s.created_at.strftime("%b %d, %Y"),
                    "timestamp": s.created_at.timestamp(),
                })
            activity_feed.sort(key=lambda x: x["timestamp"], reverse=True)
            activity_feed = activity_feed[:8]
            for item in activity_feed:
                item.pop("timestamp", None)

            records_qs = (
                AttendanceRecord.objects
                .select_related("student", "session__group", "session__course", "session__taken_by")
                .order_by("-session__session_date", "-id")[:200]
            )
            attendance_records = []
            for r in records_qs:
                attendance_records.append({
                    "id": r.id,
                    "student": r.student.username,
                    "student_name": f"{r.student.first_name} {r.student.last_name}".strip() or r.student.username,
                    "group": r.session.group.name if r.session.group else "",
                    "course": r.session.course.title if r.session.course else "",
                    "date": r.session.session_date.strftime("%Y-%m-%d"),
                    "date_display": r.session.session_date.strftime("%a, %b %d, %Y"),
                    "status": r.status,
                    "taken_by": r.session.taken_by.username if r.session.taken_by else "",
                    "month": r.session.session_date.strftime("%B"),
                })

            return Response({
                "users": {
                    "total": User.objects.count(),
                    "active": User.objects.filter(is_active=True).count(),
                    "inactive": User.objects.filter(is_active=False).count(),
                    "students": User.objects.filter(role="student").count(),
                    "instructors": User.objects.filter(role="instructor").count(),
                    "admins": User.objects.filter(role="admin").count(),
                    "new_this_week": new_users_week,
                    "new_this_month": new_users_month,
                },
                "attendance_records": attendance_records,
                "courses": {"total": Course.objects.count()},
                "lessons": {"total": Lesson.objects.count()},
                "groups": {"total": Group.objects.count()},
                "homework": {
                    "total": total_hw,
                    "submissions": total_subs,
                    "pending_grading": pending_subs,
                    "graded": graded_subs,
                    "drafts": draft_subs,
                    "avg_score": round(avg_score_all, 1),
                },
                "attendance": {
                    "sessions": AttendanceSession.objects.count(),
                    "records": total_records,
                    "rate": attendance_rate,
                    "sessions_this_week": sessions_this_week,
                },
                "finance": {
                    "total_revenue": str(total_revenue),
                    "monthly_revenue": str(monthly_revenue),
                    "weekly_revenue": str(weekly_revenue),
                    "prev_month_revenue": round(prev_month_revenue, 2),
                    "revenue_growth_pct": revenue_growth_pct,
                    "total_payments": succeeded_payments.count(),
                    "avg_payment": round(avg_payment, 2),
                    "payment_success_rate": payment_success_rate,
                    "course_conversion_rate": course_conversion_rate,
                },
                "engagement": {
                    "active_students": len(active_student_ids),
                    "total_students": total_students,
                    "engagement_rate": engagement_rate,
                },
                "top_instructors": top_instructors,
                "activity_feed": activity_feed,
                "recent_payments": [
                    {
                        "user": p.user.username,
                        "name": f"{p.user.first_name} {p.user.last_name}".strip() or p.user.username,
                        "amount": str(p.amount),
                        "date": p.created_at.strftime("%b %d, %Y"),
                    }
                    for p in recent_payments
                ],
                "recent_users": recent_users,
                "charts": {
                    "user_growth": user_growth,
                    "revenue_trend": revenue_trend,
                    "attendance_by_session": attendance_by_session,
                    "attendance_status_dist": att_status_dist,
                    "course_popularity": course_pop,
                    "payment_status_dist": payment_status_dist,
                    "revenue_by_course": revenue_by_course,
                    "daily_sessions": daily_sessions,
                    "group_attendance_breakdown": group_attendance_breakdown,
                    "attendance_rate_trend": attendance_rate_trend,
                },
            })

        elif user.role == "instructor":
            my_groups = Group.objects.filter(instructor=user)
            group_student_count = User.objects.filter(student_groups__in=my_groups).distinct().count()

            sessions = AttendanceSession.objects.filter(group__in=my_groups)
            records = AttendanceRecord.objects.filter(session__in=sessions)
            total_records = records.count()
            present_records = records.filter(status__in=["attended", "attended_online", "late"]).count()
            attendance_rate = round((present_records / total_records * 100), 1) if total_records > 0 else 0

            course_ids = my_groups.values_list("courses", flat=True).distinct()
            my_courses_count = Course.objects.filter(id__in=course_ids).count()

            my_lessons = Lesson.objects.filter(course__id__in=course_ids)
            hw_assignments = Homework.objects.filter(lesson__in=my_lessons)
            hw_submissions = HomeworkSubmission.objects.filter(homework__in=hw_assignments)

            attendance_by_session = self._attendance_status_by_session(sessions)
            att_status_dist = list(
                records.values("status").annotate(count=Count("id")).order_by("-count")
            )

            submission_trend = self._week_series(hw_submissions, "created_at", weeks=8)

            group_sizes = [
                {"name": g.name[:18], "students": g.students.count()}
                for g in my_groups[:6]
            ]

            avg_scores_by_hw = []
            for hw in hw_assignments.order_by("-id")[:8]:
                subs = hw_submissions.filter(homework=hw, score__isnull=False)
                avg = subs.aggregate(avg=Avg("score"))["avg"]
                if avg is not None:
                    avg_scores_by_hw.append({
                        "name": hw.title[:18],
                        "avg_score": round(float(avg), 1),
                        "max_score": hw.total_points,
                    })
            avg_scores_by_hw.reverse()

            group_attendance = []
            for g in my_groups[:6]:
                g_sessions = sessions.filter(group=g)
                g_records = records.filter(session__in=g_sessions)
                g_total = g_records.count()
                g_present = g_records.filter(
                    status__in=["attended", "attended_online", "late"]
                ).count()
                g_rate = round((g_present / g_total * 100), 1) if g_total > 0 else 0
                group_attendance.append({
                    "name": g.name[:18],
                    "rate": g_rate,
                    "present": g_present,
                    "total": g_total,
                })

            student_ids = User.objects.filter(student_groups__in=my_groups).distinct().values_list("id", flat=True)
            top_students = list(
                hw_submissions.filter(student__id__in=student_ids, score__isnull=False)
                .values("student__username", "student__first_name", "student__last_name")
                .annotate(avg_score=Avg("score"), total_submitted=Count("id"))
                .order_by("-avg_score")[:5]
            )
            top_students_list = [
                {
                    "username": s["student__username"],
                    "name": f"{s['student__first_name']} {s['student__last_name']}".strip() or s["student__username"],
                    "avg_score": round(float(s["avg_score"]), 1),
                    "submissions": s["total_submitted"],
                }
                for s in top_students
            ]

            recent_sess_qs = sessions.select_related("group", "course").order_by("-session_date")[:12]
            session_detail = []
            session_rate_trend = []
            for s in reversed(list(recent_sess_qs)):
                s_records = s.records.all()
                s_total = s_records.count()
                attended = s_records.filter(status="attended").count()
                online = s_records.filter(status="attended_online").count()
                late = s_records.filter(status="late").count()
                absent = s_records.filter(status="absent").count()
                excused = s_records.filter(status="excused").count()
                present = attended + online + late
                rate = round((present / s_total * 100), 1) if s_total > 0 else 0
                session_detail.append({
                    "date": s.session_date.strftime("%b %d, %Y"),
                    "short_date": s.session_date.strftime("%b %d"),
                    "group": s.group.name[:25] if s.group else "—",
                    "course": s.course.title[:25] if s.course else "—",
                    "attended": attended,
                    "online": online,
                    "late": late,
                    "excused": excused,
                    "absent": absent,
                    "total": s_total,
                    "rate": rate,
                })
                session_rate_trend.append({
                    "date": s.session_date.strftime("%b %d"),
                    "rate": rate,
                    "label": s.group.name[:18] if s.group else "—",
                })

            recent_submissions = [
                {
                    "student": s.student.username,
                    "name": f"{s.student.first_name} {s.student.last_name}".strip() or s.student.username,
                    "homework": s.homework.title[:25],
                    "date": s.created_at.strftime("%b %d, %Y"),
                    "status": s.status,
                }
                for s in hw_submissions.filter(status="submitted")
                    .select_related("student", "homework")
                    .order_by("-created_at")[:6]
            ]

            upcoming_hw = [
                {
                    "title": hw.title[:25],
                    "course": hw.lesson.course.title[:20] if hw.lesson.course else "",
                    "due_date": hw.due_date.strftime("%b %d, %Y") if hw.due_date else None,
                    "submissions": hw_submissions.filter(homework=hw).count(),
                }
                for hw in hw_assignments.filter(due_date__gte=now)
                    .select_related("lesson__course")
                    .order_by("due_date")[:5]
            ]

            return Response({
                "groups": {
                    "total": my_groups.count(),
                    "total_students": group_student_count,
                },
                "courses": {"total": my_courses_count},
                "lessons": {"total": my_lessons.count()},
                "homework": {
                    "total": hw_assignments.count(),
                    "submissions": hw_submissions.count(),
                    "pending_grading": hw_submissions.filter(status="submitted").count(),
                    "graded": hw_submissions.filter(status="graded").count(),
                },
                "attendance": {
                    "sessions": sessions.count(),
                    "records": total_records,
                    "rate": attendance_rate,
                },
                "recent_submissions": recent_submissions,
                "upcoming_homework": upcoming_hw,
                "top_students": top_students_list,
                "session_detail": session_detail,
                "attendance_records": [
                    {
                        "id": r.id,
                        "student": r.student.username,
                        "student_name": f"{r.student.first_name} {r.student.last_name}".strip() or r.student.username,
                        "group": r.session.group.name if r.session.group else "",
                        "course": r.session.course.title if r.session.course else "",
                        "date": r.session.session_date.strftime("%Y-%m-%d"),
                        "date_display": r.session.session_date.strftime("%a, %b %d, %Y"),
                        "status": r.status,
                        "month": r.session.session_date.strftime("%B"),
                    }
                    for r in records.select_related("student", "session__group", "session__course").order_by("-session__session_date", "-id")[:200]
                ],
                "charts": {
                    "attendance_by_session": attendance_by_session,
                    "attendance_status_dist": att_status_dist,
                    "submission_trend": submission_trend,
                    "group_sizes": group_sizes,
                    "avg_scores_by_hw": avg_scores_by_hw,
                    "group_attendance": group_attendance,
                    "session_rate_trend": session_rate_trend,
                },
            })

        else:           
            my_groups = user.student_groups.all()

            course_ids = my_groups.values_list("courses", flat=True).distinct()
            enrolled_courses = Course.objects.filter(id__in=course_ids)
            my_courses_count = enrolled_courses.count()

            purchased_count = CoursePurchase.objects.filter(user=user).count()

            my_records = AttendanceRecord.objects.filter(student=user)
            total_records = my_records.count()
            present_records = my_records.filter(status__in=["attended", "attended_online", "late"]).count()
            attendance_rate = round((present_records / total_records * 100), 1) if total_records > 0 else 0

            my_submissions = HomeworkSubmission.objects.filter(student=user)

            recent_records = my_records.select_related("session__group", "session__course").order_by("-session__session_date")[:5]

            att_status_dist = list(
                my_records.values("status").annotate(count=Count("id")).order_by("-count")
            )

            att_timeline = []
            recent_sessions = my_records.select_related("session").order_by("session__session_date")[:30]
            for r in recent_sessions:
                att_timeline.append({
                    "date": r.session.session_date.strftime("%b %d"),
                    "status": 1 if r.status in ["attended", "attended_online", "late"] else 0,
                    "raw_status": r.status,
                    "label": r.status.replace("_", " ").title(),
                })

            course_attendance = []
            for course in enrolled_courses[:8]:
                c_records = my_records.filter(session__course=course)
                c_total = c_records.count()
                if c_total == 0:
                    continue
                c_present = c_records.filter(
                    status__in=["attended", "attended_online", "late"]
                ).count()
                c_rate = round((c_present / c_total * 100), 1) if c_total > 0 else 0
                course_attendance.append({
                    "name": course.title[:22],
                    "rate": c_rate,
                    "present": c_present,
                    "total": c_total,
                })
            course_attendance.sort(key=lambda x: x["rate"], reverse=True)

            student_rate_trend = []
            for i in range(8):
                wk_start = (week_start - datetime.timedelta(weeks=7 - i)).date()
                wk_end = wk_start + datetime.timedelta(weeks=1)
                wk_records = my_records.filter(
                    session__session_date__gte=wk_start,
                    session__session_date__lt=wk_end,
                )
                wk_total = wk_records.count()
                wk_present = wk_records.filter(
                    status__in=["attended", "attended_online", "late"]
                ).count()
                wk_rate = round((wk_present / wk_total * 100), 1) if wk_total > 0 else None
                student_rate_trend.append({
                    "week": wk_start.strftime("%b %d"),
                    "rate": wk_rate if wk_rate is not None else 0,
                    "has_data": wk_total > 0,
                    "total": wk_total,
                })

            score_trend = []
            graded = my_submissions.filter(score__isnull=False).select_related("homework").order_by("graded_at")[:10]
            for s in graded:
                score_trend.append({
                    "name": s.homework.title[:15],
                    "score": float(s.score),
                    "total": s.homework.total_points,
                })

            course_progress = []
            for course in enrolled_courses[:6]:
                c_lessons = Lesson.objects.filter(course=course)
                c_hw = Homework.objects.filter(lesson__in=c_lessons)
                c_total = c_hw.count()
                c_done = my_submissions.filter(
                    homework__in=c_hw, status__in=["submitted", "graded"]
                ).count()
                c_avg = my_submissions.filter(
                    homework__in=c_hw, score__isnull=False
                ).aggregate(avg=Avg("score"))["avg"]
                course_progress.append({
                    "name": course.title[:22],
                    "total": c_total,
                    "done": c_done,
                    "avg_score": round(float(c_avg), 1) if c_avg else None,
                })

            all_hw = Homework.objects.filter(
                lesson__course__id__in=course_ids, due_date__gte=now
            ).select_related("lesson__course").order_by("due_date")[:5]
            upcoming_hw = []
            for hw in all_hw:
                sub = my_submissions.filter(homework=hw).first()
                upcoming_hw.append({
                    "title": hw.title[:25],
                    "course": hw.lesson.course.title[:20] if hw.lesson.course else "",
                    "due_date": hw.due_date.strftime("%b %d, %Y"),
                    "total_points": hw.total_points,
                    "status": sub.status if sub else "not_started",
                })

            score_dist = [
                {"range": "0-20%", "count": 0},
                {"range": "21-40%", "count": 0},
                {"range": "41-60%", "count": 0},
                {"range": "61-80%", "count": 0},
                {"range": "81-100%", "count": 0},
            ]
            scored_subs = my_submissions.filter(score__isnull=False).select_related("homework")
            for sub in scored_subs:
                total_pts = sub.homework.total_points or 0
                if total_pts <= 0:
                    continue
                try:
                    pct_val = (float(sub.score) / float(total_pts)) * 100.0
                except (TypeError, ValueError, ZeroDivisionError):
                    continue
                if pct_val <= 20:
                    score_dist[0]["count"] += 1
                elif pct_val <= 40:
                    score_dist[1]["count"] += 1
                elif pct_val <= 60:
                    score_dist[2]["count"] += 1
                elif pct_val <= 80:
                    score_dist[3]["count"] += 1
                else:
                    score_dist[4]["count"] += 1

            return Response({
                "groups": {"total": my_groups.count()},
                "courses": {
                    "enrolled": my_courses_count,
                    "purchased": purchased_count,
                },
                "attendance": {
                    "total": total_records,
                    "present": present_records,
                    "rate": attendance_rate,
                },
                "homework": {
                    "submitted": my_submissions.filter(status="submitted").count(),
                    "graded": my_submissions.filter(status="graded").count(),
                    "draft": my_submissions.filter(status="draft").count(),
                    "total": my_submissions.count(),
                    "average_score": float(my_submissions.filter(score__isnull=False).aggregate(avg=Avg("score"))["avg"] or 0),
                },
                "recent_attendance": [
                    {
                        "group": r.session.group.name,
                        "course": r.session.course.title if r.session.course else "General",
                        "date": r.session.session_date.strftime("%b %d, %Y"),
                        "status": r.status,
                    }
                    for r in recent_records
                ],
                "upcoming_homework": upcoming_hw,
                "course_progress": course_progress,
                "attendance_records": [
                    {
                        "id": r.id,
                        "group": r.session.group.name if r.session.group else "",
                        "course": r.session.course.title if r.session.course else "",
                        "date": r.session.session_date.strftime("%Y-%m-%d"),
                        "date_display": r.session.session_date.strftime("%a, %b %d, %Y"),
                        "status": r.status,
                        "month": r.session.session_date.strftime("%B"),
                    }
                    for r in my_records.select_related("session__group", "session__course").order_by("-session__session_date", "-id")[:200]
                ],
                "charts": {
                    "attendance_status_dist": att_status_dist,
                    "attendance_timeline": att_timeline,
                    "course_attendance": course_attendance,
                    "student_rate_trend": student_rate_trend,
                    "score_trend": score_trend,
                    "score_distribution": score_dist,
                },
            })

class RevenueTrendAPIView(APIView):
    """Filterable revenue + user-growth trend.

    Query params:
        range: one of "1d", "3d", "7d", "30d", "90d", "all" (default "7d")
    Buckets adjust automatically:
        1d   → 24 hourly buckets
        3d   → 12 six-hour buckets
        7d   → 7 daily buckets
        30d  → 30 daily buckets
        90d  → 13 weekly buckets
        all  → 12 weekly buckets
    """

    permission_classes = (permissions.IsAuthenticated, IsAdmin)

    def get(self, request):
        import datetime
        from decimal import Decimal

        from django.contrib.auth import get_user_model
        from django.db.models import Count, Sum
        from django.utils import timezone

        from apps.payments.models import Payment

        User = get_user_model()

        rng = (request.query_params.get("range") or "7d").lower()

        configs = {
            "1d": {"days": 1, "buckets": 24, "unit": "hour", "fmt": "%H:%M"},
            "3d": {"days": 3, "buckets": 12, "unit": "hour", "step_hours": 6, "fmt": "%b %d %H:%M"},
            "7d": {"days": 7, "buckets": 7, "unit": "day", "fmt": "%b %d"},
            "30d": {"days": 30, "buckets": 30, "unit": "day", "fmt": "%b %d"},
            "90d": {"days": 90, "buckets": 13, "unit": "week", "fmt": "%b %d"},
            "all": {"days": 84, "buckets": 12, "unit": "week", "fmt": "%b %d"},
        }
        cfg = configs.get(rng, configs["7d"])

        now = timezone.now()
        start = now - datetime.timedelta(days=cfg["days"])

        payments = list(
            Payment.objects.filter(status="succeeded", created_at__gte=start)
            .values_list("created_at", "amount")
        )
        new_users = list(
            User.objects.filter(date_joined__gte=start).values_list("date_joined", flat=True)
        )

        buckets = []
        if cfg["unit"] == "hour":
            step = datetime.timedelta(hours=cfg.get("step_hours", 1))
            cursor = start
            while cursor < now:
                end = cursor + step
                buckets.append({"start": cursor, "end": end, "label": cursor.strftime(cfg["fmt"])})
                cursor = end
        elif cfg["unit"] == "day":
            day_start = start.replace(hour=0, minute=0, second=0, microsecond=0)
            for i in range(cfg["buckets"]):
                cursor = day_start + datetime.timedelta(days=i)
                end = cursor + datetime.timedelta(days=1)
                buckets.append({"start": cursor, "end": end, "label": cursor.strftime(cfg["fmt"])})
        else:        
            week_start = (start - datetime.timedelta(days=start.weekday())).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            for i in range(cfg["buckets"]):
                cursor = week_start + datetime.timedelta(weeks=i)
                end = cursor + datetime.timedelta(weeks=1)
                buckets.append({"start": cursor, "end": end, "label": cursor.strftime(cfg["fmt"])})

        result = []
        for b in buckets:
            rev = sum(
                float(amt) for ts, amt in payments if b["start"] <= ts < b["end"]
            )
            users_count = sum(1 for ts in new_users if b["start"] <= ts < b["end"])
            result.append({
                "period": b["label"],
                "revenue": round(rev, 2),
                "users": users_count,
            })

        total_revenue = sum(p["revenue"] for p in result)
        total_users = sum(p["users"] for p in result)

        return Response({
            "range": rng,
            "data": result,
            "summary": {
                "total_revenue": round(total_revenue, 2),
                "total_users": total_users,
                "data_points": len(result),
            },
        })

class ExportExcelAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsAdmin)

    def get(self, request):
        from django.contrib.auth import get_user_model
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        from apps.attendance.models import AttendanceRecord, AttendanceSession
        from apps.courses.models import Course
        from apps.groups.models import Group
        from apps.homework.models import Homework, HomeworkSubmission
        from apps.lessons.models import Lesson
        from apps.payments.models import CoursePurchase, Payment

        User = get_user_model()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")

        def _write_sheet(wb, title, headers, rows):
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for row in rows:
                ws.append(row)
            for col_idx, _ in enumerate(headers, 1):
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                    default=0,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        wb = Workbook()
        wb.remove(wb.active)

        _write_sheet(
            wb, "Users",
            ["Email", "Username", "First Name", "Last Name", "Role", "Active", "Date Joined"],
            [
                [u.email, u.username, u.first_name, u.last_name, u.role, u.is_active, u.date_joined.strftime("%Y-%m-%d %H:%M")]
                for u in User.objects.all().order_by("id")
            ],
        )

        _write_sheet(
            wb, "Courses",
            ["Title", "Description", "Price", "Created At"],
            [
                [c.title, c.description, float(c.price), c.created_at.strftime("%Y-%m-%d %H:%M")]
                for c in Course.objects.all().order_by("id")
            ],
        )

        groups_qs = Group.objects.select_related("instructor").prefetch_related("students", "courses").order_by("id")
        _write_sheet(
            wb, "Groups",
            ["Name", "Description", "Instructor", "Student Count", "Course Count", "Created At"],
            [
                [g.name, g.description, g.instructor.username if g.instructor else "", g.students.count(), g.courses.count(), g.created_at.strftime("%Y-%m-%d %H:%M")]
                for g in groups_qs
            ],
        )

        lessons_qs = Lesson.objects.select_related("course", "user").order_by("id")
        _write_sheet(
            wb, "Lessons",
            ["Title", "Course", "Instructor", "Created At"],
            [
                [l.title, l.course.title, l.user.username, l.created_at.strftime("%Y-%m-%d %H:%M")]
                for l in lessons_qs
            ],
        )

        payments_qs = Payment.objects.select_related("user").order_by("id")
        _write_sheet(
            wb, "Payments",
            ["User", "Amount", "Currency", "Status", "Created At"],
            [
                [p.user.username, float(p.amount), p.currency, p.status, p.created_at.strftime("%Y-%m-%d %H:%M")]
                for p in payments_qs
            ],
        )

        purchases_qs = CoursePurchase.objects.select_related("user", "course").order_by("id")
        _write_sheet(
            wb, "Course Purchases",
            ["User", "Course", "Amount", "Created At"],
            [
                [cp.user.username, cp.course.title, float(cp.amount), cp.created_at.strftime("%Y-%m-%d %H:%M")]
                for cp in purchases_qs
            ],
        )

        sessions_qs = AttendanceSession.objects.select_related("group", "course", "taken_by").order_by("id")
        _write_sheet(
            wb, "Attendance Sessions",
            ["Group", "Course", "Taken By", "Session Date"],
            [
                [s.group.name, s.course.title if s.course else "", s.taken_by.username if s.taken_by else "", s.session_date.strftime("%Y-%m-%d")]
                for s in sessions_qs
            ],
        )

        records_qs = AttendanceRecord.objects.select_related("student", "session__group", "session__course").order_by("id")
        _write_sheet(
            wb, "Attendance Records",
            ["Student", "Group", "Course", "Session Date", "Status"],
            [
                [r.student.username, r.session.group.name, r.session.course.title if r.session.course else "", r.session.session_date.strftime("%Y-%m-%d"), r.status]
                for r in records_qs
            ],
        )

        hw_qs = Homework.objects.select_related("lesson__course", "created_by").order_by("id")
        _write_sheet(
            wb, "Homework",
            ["Title", "Lesson", "Course", "Total Points", "Due Date", "Created By"],
            [
                [h.title, h.lesson.title, h.lesson.course.title, h.total_points, h.due_date.strftime("%Y-%m-%d %H:%M") if h.due_date else "", h.created_by.username if h.created_by else ""]
                for h in hw_qs
            ],
        )

        subs_qs = HomeworkSubmission.objects.select_related("student", "homework", "graded_by").order_by("id")
        _write_sheet(
            wb, "Homework Submissions",
            ["Student", "Homework", "Status", "Score", "Graded By", "Submitted At"],
            [
                [s.student.username, s.homework.title, s.status, float(s.score) if s.score is not None else "", s.graded_by.username if s.graded_by else "", s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else ""]
                for s in subs_qs
            ],
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="lms_export.xlsx"'
        return response

class AIChatAPIView(APIView):
    """Gemini-powered tutor chatbot for authenticated students.

    POST body:
        messages: [{role: "user"|"model", content: "..."}, ...]

    Returns:
        { reply: str, model: str }
    """

    permission_classes = (permissions.IsAuthenticated,)

    GEMINI_MODEL = "gemini-2.0-flash"
    GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    MAX_HISTORY = 20
    MAX_INPUT_CHARS = 4000

    def _build_context(self, user):
        """Build a compact, student-specific context block for the system prompt."""
        from datetime import timedelta
        from django.db.models import Avg
        from django.utils import timezone

        from apps.attendance.models import AttendanceRecord
        from apps.courses.models import Course
        from apps.homework.models import Homework, HomeworkSubmission
        from apps.lessons.models import Lesson

        ctx_lines = []

        if user.role == "student":
            my_groups = user.student_groups.all()
            course_ids = list(my_groups.values_list("courses", flat=True).distinct())
            courses = Course.objects.filter(id__in=course_ids)
            if courses.exists():
                ctx_lines.append(
                    "Enrolled courses: " + ", ".join(c.title for c in courses[:8])
                )

            lessons = Lesson.objects.filter(course__id__in=course_ids).order_by("-created_at")[:6]
            if lessons:
                ctx_lines.append(
                    "Recent lessons: " + ", ".join(f"{l.title} ({l.course.title})" for l in lessons)
                )

            now = timezone.now()
            upcoming = Homework.objects.filter(
                lesson__course__id__in=course_ids, due_date__gte=now
            ).select_related("lesson__course").order_by("due_date")[:5]
            if upcoming:
                hw_lines = []
                for hw in upcoming:
                    sub = HomeworkSubmission.objects.filter(homework=hw, student=user).first()
                    status_str = sub.status if sub else "not started"
                    hw_lines.append(
                        f"\"{hw.title}\" ({hw.lesson.course.title}, due {hw.due_date.strftime('%b %d')}, {hw.total_points}pts, status: {status_str})"
                    )
                ctx_lines.append("Upcoming homework: " + "; ".join(hw_lines))

            my_subs = HomeworkSubmission.objects.filter(student=user, score__isnull=False)
            avg_score = my_subs.aggregate(avg=Avg("score"))["avg"]
            if avg_score is not None:
                ctx_lines.append(f"Average homework score: {round(float(avg_score), 1)} pts")

            recent_records = AttendanceRecord.objects.filter(
                student=user, session__session_date__gte=(now - timedelta(days=60)).date()
            )
            tot = recent_records.count()
            present = recent_records.filter(
                status__in=["attended", "attended_online", "late"]
            ).count()
            if tot > 0:
                rate = round((present / tot) * 100, 1)
                ctx_lines.append(f"Attendance rate (last 60 days): {rate}% ({present}/{tot})")

        full_name = f"{user.first_name} {user.last_name}".strip() or user.username
        return full_name, ctx_lines

    def post(self, request):
        import json
        from urllib import error as urlerr, request as urlreq

        from django.conf import settings

        api_key = getattr(settings, "GEMINI_API_KEY", None) or os.environ.get("GEMINI_API_KEY")
        if not api_key:
            return Response(
                {"detail": "AI chat is not configured. Server admin must set GEMINI_API_KEY."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        messages = request.data.get("messages") or []
        if not isinstance(messages, list) or not messages:
            return Response(
                {"detail": "messages must be a non-empty list."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        messages = messages[-self.MAX_HISTORY:]
        for m in messages:
            content = m.get("content", "")
            if not isinstance(content, str):
                return Response({"detail": "Each message needs a string content."}, status=400)
            if len(content) > self.MAX_INPUT_CHARS:
                return Response(
                    {"detail": f"Message too long (max {self.MAX_INPUT_CHARS} chars)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        full_name, ctx_lines = self._build_context(request.user)

        system_prompt = (
            f"You are an AI tutor inside an online learning management system (LMS). "
            f"You're chatting with {full_name} (role: {request.user.role}). "
            "Be friendly, concise, and educational. When explaining concepts, use clear examples. "
            "If the student asks about their progress, courses, lessons, or homework, "
            "use the context below. If the answer is not in the context, say so honestly. "
            "Do not invent grades, dates, or facts about the student. "
            "Format code with markdown code blocks. Keep replies under ~250 words unless asked for detail.\n\n"
        )
        if ctx_lines:
            system_prompt += "STUDENT CONTEXT:\n- " + "\n- ".join(ctx_lines)
        else:
            system_prompt += "STUDENT CONTEXT: (none available)"

        contents = []
        for m in messages:
            role = m.get("role", "user")
            gemini_role = "model" if role in ("model", "assistant") else "user"
            contents.append({
                "role": gemini_role,
                "parts": [{"text": str(m.get("content", ""))}],
            })

        body = {
            "system_instruction": {"parts": [{"text": system_prompt}]},
            "contents": contents,
            "generationConfig": {
                "temperature": 0.7,
                "maxOutputTokens": 1024,
                "topP": 0.95,
            },
            "safetySettings": [
                {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_ONLY_HIGH"},
                {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_ONLY_HIGH"},
                {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_ONLY_HIGH"},
                {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_ONLY_HIGH"},
            ],
        }

        url = self.GEMINI_URL.format(model=self.GEMINI_MODEL) + f"?key={api_key}"
        req = urlreq.Request(
            url,
            data=json.dumps(body).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )

        try:
            with urlreq.urlopen(req, timeout=30) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
        except urlerr.HTTPError as e:
            try:
                err_body = json.loads(e.read().decode("utf-8"))
                detail = err_body.get("error", {}).get("message", str(e))
            except Exception:
                detail = str(e)
            return Response(
                {"detail": f"AI provider error: {detail}"},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except Exception as e:
            return Response(
                {"detail": f"AI request failed: {e}"},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        candidates = payload.get("candidates") or []
        if not candidates:
            block = payload.get("promptFeedback", {}).get("blockReason")
            return Response(
                {"detail": f"AI returned no candidates" + (f" (blocked: {block})" if block else "")},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        parts = candidates[0].get("content", {}).get("parts", [])
        reply_text = "".join(p.get("text", "") for p in parts).strip()
        if not reply_text:
            reply_text = "(empty response)"

        return Response({"reply": reply_text, "model": self.GEMINI_MODEL})
